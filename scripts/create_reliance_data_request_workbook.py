from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("outputs/019feac5-aea6-7e13-b42f-878ff63c84cd")
OUTPUT_FILE = OUTPUT_DIR / "reliance_geology_data_request_pack.xlsx"


COLORS = {
    "title": "17324D",
    "header": "1F4E5F",
    "header_alt": "385723",
    "light": "EAF3F8",
    "light_green": "E2F0D9",
    "light_yellow": "FFF2CC",
    "light_red": "FCE4D6",
    "border": "D9E2F3",
    "white": "FFFFFF",
    "text": "1F2933",
}


def clean_table_name(name: str) -> str:
    return "".join(ch if ch.isalnum() else "_" for ch in name).strip("_")[:250]


def set_common_style(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def add_table(ws, data: list[list], table_name: str, header_color: str = COLORS["header"]) -> None:
    if not data:
        return
    ws.append(data[0])
    for row in data[1:]:
        ws.append(row)

    max_row = ws.max_row
    max_col = ws.max_column
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    header_fill = PatternFill("solid", fgColor=header_color)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color=COLORS["white"], bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    band_fill = PatternFill("solid", fgColor="F7FAFC")
    for row_number in range(2, max_row + 1):
        if row_number % 2 == 0:
            for cell in ws[row_number]:
                cell.fill = band_fill


def autofit(ws, min_width: int = 10, max_width: int = 52) -> None:
    for column in ws.columns:
        letter = column[0].column_letter
        max_len = 0
        for cell in column:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            max_len = max(max_len, min(len(text), max_width))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 32 if row == 1 else 48


def apply_light_borders(ws) -> None:
    side = Side(style="thin", color=COLORS["border"])
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=side)


def add_validation(ws, cell_range: str, values: list[str]) -> None:
    csv_values = ",".join(values)
    dv = DataValidation(type="list", formula1=f'"{csv_values}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_comments(ws, comments: dict[str, str]) -> None:
    for cell_ref, text in comments.items():
        ws[cell_ref].comment = Comment(text, "GeoWorkbench")


def add_priority_formatting(ws, priority_col: str = "D", start_row: int = 2, end_row: int = 500) -> None:
    ws.conditional_formatting.add(
        f"{priority_col}{start_row}:{priority_col}{end_row}",
        FormulaRule(
            formula=[f'{priority_col}{start_row}="P0 - Critical"'],
            fill=PatternFill("solid", fgColor=COLORS["light_red"]),
        ),
    )
    ws.conditional_formatting.add(
        f"{priority_col}{start_row}:{priority_col}{end_row}",
        FormulaRule(
            formula=[f'{priority_col}{start_row}="P1 - Needed"'],
            fill=PatternFill("solid", fgColor=COLORS["light_yellow"]),
        ),
    )


def create_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "00_ReadMe"
    ws.sheet_view.showGridLines = False
    rows = [
        ["Reliance Coal Geology Data Request Pack", ""],
        ["Prepared for", "GeoWorkbench web + mobile assistant for coal geologists"],
        ["Purpose", "Request the datasets needed to mature import, log visualization, mobile field capture, analytics, AI assistance, seam correlation, and export workflows."],
        ["NDA context", "Use this for NDA-covered Reliance data transfer. Do not commit confidential customer or mine data to the public repository."],
        ["Recommended first delivery", "10-20 complete boreholes from one block, with matching collars, lithology intervals, seams, LAS logs, core image mapping, coal quality samples, dictionaries, and final approved corrections."],
        ["Recommended AI/correlation delivery", "50+ historical approved boreholes for statistical rules and correlation QA; 100-300+ depth-indexed boreholes with paired logs and approved labels for robust supervised lithology/seam models."],
        ["Workbook structure", "Sheets 03-15 are request templates. Reliance can either fill these sheets directly or provide existing files with a mapping to the same fields."],
        ["Source principle", "Preserve raw files, normalized tables, interpreted/corrected layers, AI suggestions, and geologist feedback separately."],
        ["Preferred file formats", "XLSX/CSV for tabular logs, LAS for geophysical curves, GeoPackage/SHP/GeoJSON for spatial data, original image files plus manifests for core photos, PDF only as evidence or fallback."],
        ["Important note", "The workbook was generated using the repo's openpyxl-compatible path because the preferred spreadsheet artifact runtime was unavailable in this session."],
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(size=18, bold=True, color=COLORS["white"])
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["title"])
    ws["B1"].fill = PatternFill("solid", fgColor=COLORS["title"])
    for cell in ws["A"]:
        cell.font = Font(bold=True, color=COLORS["text"])
        cell.fill = PatternFill("solid", fgColor=COLORS["light"])
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100
    apply_light_borders(ws)


def create_dataset_request(wb: Workbook) -> None:
    ws = wb.create_sheet("01_Dataset_Request")
    rows = [
        [
            "Dataset_ID",
            "Dataset",
            "Why_we_need_it",
            "Priority",
            "Product_modules_enabled",
            "Preferred_format",
            "Minimum_fields_or_files",
            "Nice_to_have",
            "AI_or_correlation_value",
            "Initial_volume_request",
            "Owner_to_identify",
            "Sensitivity",
            "Acceptance_checks",
            "Reference_URL",
        ],
        ["REL-001", "Project/block and borehole master", "Creates the central identity layer for every import, mobile submission, map, correlation view, and export.", "P0 - Critical", "Import Center, Workbench, Mobile, Correlation, Export", "XLSX/CSV plus GIS layer", "project/block, borehole id, collar coordinates, RL/elevation, total depth, status, dates, drilling method, coordinate system", "deviation survey, casing, diameter, contractor, planned vs final depth", "Required for every join, spatial trend, RL alignment, and duplicate detection.", "All boreholes in selected pilot block; then historical blocks", "Reliance geology/data admin", "High", "Unique borehole IDs; CRS declared; total depth matches interval/log coverage.", "https://mecl.co.in/ContentPageMecl.aspx?Antispam=5986bcdc-bac7-4286-8f12-6f7e9bac0700&ControlID=4&Lng=EN&MyAntispam=a3782f6e-6c86-401c-b726-e2951ccb487d&page=mineral-exploration"],
        ["REL-002", "Raw descriptive lithology/geointerval logs", "Core data for the log widget, interval editing, validation rules, AI review, and export-ready corrected logs.", "P0 - Critical", "Log Widget, Interval Details, Validation, AI Workflow, Export", "Original XLSX plus normalized CSV", "from depth, to/thickness, lithology source text/code, recovery, RQD, color, structural features, core dip, seam, remarks, source row", "weathering, grain size, hardness, bedding, fractures, fossils, sample links, correction stage", "Supervised labels for lithology prediction; sequence context for correlation and anomaly detection.", "10-20 complete boreholes for pilot; 100+ approved boreholes for ML", "Site geologists and central geology team", "High", "No unreviewed gaps/overlaps; source row preserved; unit format known.", "https://www.acarp.com.au/abstracts.aspx?repId=C21003"],
        ["REL-003", "Drilling run and recovery data", "Separates drilling/core recovery intervals from lithology intervals, supporting recovery/RQD QA and field-to-central reconciliation.", "P0 - Critical", "Mobile, Workbench, Quantitative Tracks, Validation", "XLSX/CSV", "run from/to, run length, recovered length, recovery %, bit/core size, run date, driller notes", "core loss reason, drilling fluid, casing, breakdown/caving remarks", "Improves anomaly explanations where core loss or washout affects interpretation.", "Same boreholes as lithology logs", "Drilling contractor/site geologist", "High", "Recovery does not exceed run length unless explicitly explained.", "https://www.ausimm.com/globalassets/coallog/coallog-v2.0-training-manual.pdf"],
        ["REL-004", "Seam intervals, markers, and correlation picks", "Needed for seam track, multi-borehole correlation, seam continuity checks, and downstream modelling exports.", "P0 - Critical", "Seam Track, Correlation Workspace, AI Assistant, Export", "XLSX/CSV; modelling package exports if available", "borehole id, seam name, top depth, base depth, thickness, partings, confidence, correlation status", "RL top/base, roof/floor lithology, split/merge flags, neighbouring evidence, fault notes", "Primary target for seam matching, missing seam detection, thickness trend analytics.", "All pilot boreholes with final approved picks; historical picks for at least one mature block", "Central geologists/modelling team", "High", "Seam aliases reconciled; top/base within lithology and total depth.", "https://www.ausimm.com/globalassets/coallog/coallog-v2.0-training-manual.pdf"],
        ["REL-005", "Geophysical logs", "Raw depth-indexed curves enable curve display, curve/lithology mismatch checks, coal candidate detection, and depth adjustment evidence.", "P0 - Critical", "Curve Track, Curve Catalog, Validation, AI Workflow, Correlation, Export LAS", "LAS 2.0/3.0 preferred; CSV fallback; PDF only fallback", "raw LAS files with WELL, CURVE, ASCII sections; gamma, density, resistivity, caliper, SP, sonic, deviation where available", "tool metadata, calibration, logging contractor, filters, runs, depth shifts, null values, interpreted tops", "Core evidence for lithology/seam prediction and cross-borehole curve signature matching.", "At least one matched LAS per pilot borehole; all historical LAS for AI", "Geophysical logging contractor/geology team", "High", "Borehole ID matches collar/log; depth coverage and null handling documented.", "https://www.usgs.gov/programs/national-geological-and-geophysical-data-preservation-program/las-format"],
        ["REL-006", "Core image files and depth mapping", "Links visual evidence to intervals and enables future computer vision review of coal, fractures, missing core, labels, and image-depth conflicts.", "P0 - Critical", "Image Track, Interval Details, Mobile Uploads, AI Evidence", "Original JPG/PNG plus image manifest CSV/XLSX", "borehole id, box number, file name, from/to depth, lane/row count, image order, capture date, depth labels", "cropped strip assets, OCR labels, annotations, lighting/scale details, geologist-confirmed mapping", "Needed for image AI and evidence-linked summaries; prevents unreliable visual suggestions.", "Full image folder for 10-20 pilot boreholes; annotated crops later", "Core shed/site geology team", "High", "Every image has box/depth mapping or explicit pending status.", "https://www.ausimm.com/globalassets/coallog/coallog-v2.0-training-manual.pdf"],
        ["REL-007", "Coal quality assay and sample dispatch data", "Matures analytics beyond logs: ash/GCV/moisture trends, seam quality summaries, resource/report support, and future quality prediction.", "P1 - Needed", "Analytics, AI Assistant, Seam Summary, Reporting, Export", "XLSX/CSV plus lab PDFs if available", "sample id, borehole id, from/to depth, seam, sample type, basis, lab, ash, moisture, volatile matter, fixed carbon, sulfur, GCV", "washability, density fractions, HGI, forms of sulfur, lab QA/QC, repeat samples, UHV/NCV", "Enables seam quality estimation and outlier/reconciliation analytics.", "All pilot seam samples; historical lab database for model learning", "Quality/laboratory team", "High", "Sample depth links to intervals/seams; basis units stated.", "https://mecl.co.in/ContentPageMecl.aspx?Antispam=5986bcdc-bac7-4286-8f12-6f7e9bac0700&ControlID=4&Lng=EN&MyAntispam=a3782f6e-6c86-401c-b726-e2951ccb487d&page=mineral-exploration"],
        ["REL-008", "Approved geological reports, sections, contour plans, and seam folios", "Provides source-linked knowledge assistant content and validates what the final product must export/report.", "P1 - Needed", "AI Assistant, Reporting, Correlation, Knowledge Search", "PDF/DOCX plus source tables where possible", "final geological reports, lithologs, seam correlation charts, structure/floor contours, maps, annexures", "editable source files, report templates, cross-section data, plate metadata", "RAG/knowledge assistant corpus with citations; benchmark for report drafting.", "2-5 final reports for pilot block plus historical report library", "Central geology/document control", "High", "Reports can be tied to block, borehole, seam, and date/version.", "https://www.pib.gov.in/PressReleasePage.aspx?PRID=1744880"],
        ["REL-009", "Spatial/GIS layers and coordinate reference systems", "Needed for map context, mobile offline capture, collar validation, section line selection, and spatial analytics.", "P1 - Needed", "Mobile, Maps, Correlation, Analytics", "GeoPackage preferred; SHP/GeoJSON/DXF/DWG accepted", "block boundary, lease, collars, topography, faults, lineaments, roads, pits, villages, drainage, CRS/EPSG", "DEM, satellite/drone imagery, old plan scans, grid/local mine coordinate transform", "Supports spatial trend features, section creation, and offline mobile packages.", "Pilot block full GIS package", "GIS/survey team", "High", "All layers have CRS; local grid transforms documented.", "https://www.ogc.org/standards/geopackage/"],
        ["REL-010", "Geotechnical, hydrogeology, gas, and runtime field parameters", "Matures the mobile workflow and future safety/geotechnical analytics around RQD, defects, water, gas, roof/floor and drilling conditions.", "P1 - Needed", "Mobile, Analytics, AI Assistant, Validation", "XLSX/CSV; mobile form definitions", "RQD, fractures/defects, point load/UCS, water levels/flows, gas/desorption if available, runtime drilling parameters", "ROP, torque, bit depth, hole depth, flush/washout notes, geologist hazard observations", "Improves anomaly explanations and future roof/floor/risk analytics.", "Pilot boreholes plus active field form samples", "Geotech/hydro/site operations", "High", "Units and measurement methods declared; timestamp/depth tied.", "https://www.ausimm.com/globalassets/coallog/coallog-v2.0-training-manual.pdf"],
        ["REL-011", "Dictionaries, aliases, units, and approval workflow rules", "Prevents silent bad normalization and makes import/export repeatable across Reliance templates and local geology vocabulary.", "P0 - Critical", "Import Profiles, Log Widget, AI, Validation, Export", "XLSX/CSV", "lithology codes, seam aliases, color/patterns, unit conventions, sample types, correction stages, approval statuses", "CoalLog mapping, Minex field mapping, role permissions, naming governance", "Stabilizes labels for AI and correlation; records customer-specific ontology.", "Current dictionaries plus examples of unknown/legacy codes", "Chief geologist/data steward", "High", "Every code has source label, approved label, status, and owner.", "https://www.acarp.com.au/abstracts.aspx?repId=C21003"],
        ["REL-012", "Import/export target templates and downstream system examples", "Needed to finalize Import Center and Export Center for Reliance rather than only generic Excel/CSV/LAS outputs.", "P0 - Critical", "Import Center, Export Center, Audit, Downstream Integrations", "Original templates, sample exports/imports", "current input workbooks, Minex/Surpac/other import templates, required columns, naming conventions, null handling", "API/database access details, validation reports, maker-checker approval requirements", "Ensures model-ready export and avoids manual column reshaping after approval.", "One exact template per downstream destination", "Modelling/planning/data admin", "High", "A generated export can be re-imported into target software without manual edits.", "https://www.usgs.gov/programs/national-geological-and-geophysical-data-preservation-program/las-format"],
        ["REL-013", "Human correction and feedback history", "Turns the app into a learning system: accepted/rejected suggestions, edited intervals, reasons, and final approvals become training labels.", "P1 - Needed", "AI Workflow, Audit, Learning Loop, Analytics", "XLSX/CSV/database extract", "entity id, before/after values, changed by, timestamp, reason, source evidence, approval status", "review comments, confidence, rejected AI rationale, second reviewer", "Most valuable customer-specific AI dataset after raw logs.", "Start capturing immediately; historical correction logs if available", "Central geology/admin", "High", "Corrections are linked to source interval and final approved stage.", "https://www.acarp.com.au/abstracts.aspx?repId=C21003"],
        ["REL-014", "Production/model reconciliation datasets", "Later-stage maturity: compare geological model expectations with mined/production/quality outcomes.", "P2 - Mature solution", "Analytics, Knowledge Assistant, Reconciliation, Executive Dashboards", "CSV/XLSX/database exports", "model seam/block, planned quality/thickness, mined actuals, dispatch/stockpile quality, date/bench/block", "fleet/dispatch linkage, weighbridge, washery, stockpile surveys", "Enables business-facing analytics and learning from actual outcomes.", "Only after borehole/geology pilot is stable", "Operations/planning/quality", "Very High", "Identifiers align across model, production, quality, and dispatch.", "https://mecl.co.in/ContentPageMecl.aspx?Antispam=5986bcdc-bac7-4286-8f12-6f7e9bac0700&ControlID=4&Lng=EN&MyAntispam=a3782f6e-6c86-401c-b726-e2951ccb487d&page=mineral-exploration"],
    ]
    add_table(ws, rows, "Dataset_Request")
    set_common_style(ws)
    autofit(ws)
    apply_light_borders(ws)
    add_validation(ws, "D2:D200", ["P0 - Critical", "P1 - Needed", "P2 - Mature solution"])
    add_validation(ws, "L2:L200", ["Medium", "High", "Very High"])
    add_priority_formatting(ws, "D", 2, 200)


def create_delivery_plan(wb: Workbook) -> None:
    ws = wb.create_sheet("02_Delivery_Plan")
    rows = [
        ["Wave", "Timing", "Datasets", "Goal", "Reliance_action", "GeoWorkbench_action", "Exit_criteria"],
        ["Wave 0 - Alignment", "Week 0", "Dictionaries, current templates, software/export examples, data owner list", "Agree the exact field vocabulary and target workflow.", "Nominate data owners and provide non-confidential sample schemas.", "Map to canonical model and finalize secure transfer checklist.", "Template variants and field dictionary are approved."],
        ["Wave 1 - Pilot block", "Weeks 1-2", "10-20 complete matched boreholes: collars, lithology, seams, LAS, core image manifest, coal quality, final reports", "Replace random/demo data with real, matched, reviewable Reliance data.", "Deliver pilot package under NDA with file inventory.", "Import, profile, validate, and show gaps/conflicts.", "Workbench shows real logs, curves, images, seams, source traceability."],
        ["Wave 2 - Correlation dataset", "Weeks 3-5", "50+ historical approved boreholes from one or more blocks, correlation charts, marker tables, GIS/fault layers", "Mature correlation and anomaly rules with real stratigraphic variation.", "Provide historical final approved picks and maps.", "Build correlation QA and seam trend analytics.", "Correlation workspace can compare boreholes and flag seam inconsistencies."],
        ["Wave 3 - AI learning corpus", "Weeks 5-10", "100-300+ boreholes with approved labels, LAS curves, corrections, quality samples, image annotations if available", "Train and evaluate customer-specific lithology/seam/quality models.", "Provide approved historical labels and correction history.", "Build training splits, baselines, explainability, confidence scoring.", "Model suggestions are evidence-linked and reviewed by geologists."],
        ["Wave 4 - Enterprise maturity", "After pilot", "Production reconciliation, planning/model exports, document corpus, operational systems", "Move from borehole workbench to mine intelligence layer.", "Prioritize integrations and governance requirements.", "Implement connectors, dashboards, approval workflows.", "Approved data flows to downstream systems without manual reshaping."],
    ]
    add_table(ws, rows, "Delivery_Plan", COLORS["header_alt"])
    set_common_style(ws)
    autofit(ws)
    apply_light_borders(ws)


def create_borehole_master(wb: Workbook) -> None:
    ws = wb.create_sheet("03_Borehole_Master")
    headers = [
        "project_code",
        "block_name",
        "coalfield_or_basin",
        "site_or_sector",
        "borehole_id",
        "alternate_borehole_id",
        "borehole_title",
        "state",
        "country",
        "coordinate_system",
        "epsg_code",
        "local_grid_name",
        "easting",
        "northing",
        "latitude",
        "longitude",
        "elevation_rl_m",
        "collar_survey_method",
        "planned_depth_m",
        "final_depth_m",
        "start_date",
        "completion_date",
        "status",
        "drilling_method",
        "hole_diameter",
        "casing_depth_m",
        "water_level_m",
        "inclination_deg",
        "azimuth_deg",
        "drilling_contractor",
        "site_geologist",
        "central_reviewer",
        "source_file_inventory",
        "remarks",
    ]
    rows = [headers] + [["" for _ in headers] for _ in range(100)]
    add_table(ws, rows, "Borehole_Master")
    set_common_style(ws)
    add_validation(ws, "W2:W101", ["planned", "drilling", "logging_in_progress", "closed", "approved", "archived"])
    add_validation(ws, "X2:X101", ["core", "non_core", "RC", "chip", "mixed", "unknown"])
    add_comments(
        ws,
        {
            "E1": "Primary key used across every sheet.",
            "J1": "State the coordinate reference system or local mine grid.",
            "K1": "Use EPSG where possible. If local grid is used, provide transform details in GIS sheet.",
            "T1": "Should match deepest interval and geophysical log coverage unless intentionally partial.",
        },
    )
    autofit(ws, max_width=34)
    apply_light_borders(ws)


def create_geointervals(wb: Workbook) -> None:
    ws = wb.create_sheet("04_GeoIntervals")
    headers = [
        "project_code",
        "borehole_id",
        "interval_id",
        "source_file",
        "source_sheet",
        "source_row",
        "from_depth_m",
        "to_depth_m",
        "thickness_m",
        "drilling_run_from_m",
        "drilling_run_to_m",
        "run_recovery_m",
        "run_recovery_pct",
        "lithology_source_text",
        "lithology_code",
        "lithology_label",
        "lithology_dictionary_status",
        "logged_color",
        "grain_size",
        "weathering",
        "hardness",
        "structural_features",
        "core_dip_deg",
        "rqd_piece_lengths",
        "rqd_pct",
        "recovery_m",
        "recovery_pct",
        "seam_name",
        "roof_floor_role",
        "sample_id",
        "image_box",
        "image_file",
        "remarks",
        "logged_by",
        "logged_at",
        "correction_stage",
        "approval_status",
        "qa_flags",
    ]
    rows = [headers]
    for i in range(2, 502):
        row = ["" for _ in headers]
        row[8] = f"=IF(OR(G{i}=\"\",H{i}=\"\"),\"\",H{i}-G{i})"
        row[12] = f"=IF(OR(J{i}=\"\",K{i}=\"\",L{i}=\"\"),\"\",L{i}/(K{i}-J{i}))"
        row[26] = f"=IF(OR(I{i}=\"\",Z{i}=\"\"),\"\",Z{i}/I{i})"
        rows.append(row)
    add_table(ws, rows, "GeoIntervals")
    set_common_style(ws)
    add_validation(ws, "Q2:Q501", ["approved", "alias_review", "unknown", "do_not_map"])
    add_validation(ws, "AJ2:AJ501", ["raw_imported", "field_submitted", "central_corrected", "approved_final"])
    add_validation(ws, "AK2:AK501", ["draft", "needs_review", "approved", "rejected"])
    for col in ["G", "H", "I", "J", "K", "L", "M", "W", "Y", "Z", "AA"]:
        for cell in ws[f"{col}2:{col}501"]:
            cell[0].number_format = "0.00"
    add_comments(
        ws,
        {
            "O1": "Store the source/local code exactly; map approved labels in Dictionaries sheet.",
            "Q1": "Do not silently normalize unknown lithology codes. Mark alias_review or unknown.",
            "Y1": "Use percent value 0-100. The app can normalize to fraction internally.",
            "AJ1": "Separate raw field, central corrected, and approved final stages.",
        },
    )
    autofit(ws, max_width=32)
    apply_light_borders(ws)


def create_curve_metadata(wb: Workbook) -> None:
    ws = wb.create_sheet("05_Curve_Metadata")
    headers = [
        "project_code",
        "borehole_id",
        "source_las_file",
        "source_pdf_or_csv_if_no_las",
        "logging_contractor",
        "logging_date",
        "log_run_number",
        "depth_reference",
        "start_depth_m",
        "stop_depth_m",
        "step_m",
        "null_value",
        "curve_mnemonic",
        "canonical_curve_key",
        "curve_description",
        "unit",
        "tool_name",
        "calibration_info",
        "raw_or_processed",
        "filter_or_smoothing",
        "depth_shift_applied_m",
        "sample_count",
        "coverage_pct",
        "min_value",
        "max_value",
        "quality_status",
        "comments",
    ]
    rows = [headers] + [["" for _ in headers] for _ in range(200)]
    add_table(ws, rows, "Curve_Metadata")
    set_common_style(ws)
    add_validation(ws, "N2:N201", ["gamma", "density", "resistivity", "caliper", "sp", "sonic", "inclination", "azimuth", "other"])
    add_validation(ws, "S2:S201", ["raw", "processed", "interpreted", "digitized_pdf"])
    add_validation(ws, "Z2:Z201", ["usable", "partial", "needs_review", "rejected"])
    add_comments(
        ws,
        {
            "C1": "Preferred source is original LAS 2.0/3.0 where available.",
            "M1": "Keep original LAS mnemonic, e.g. NGAM, GR, RHOB, DENS, 16N, 64N, RES, CAL, SP.",
            "N1": "Canonical key used by GeoWorkbench curve track and validation rules.",
            "U1": "Document any depth correction before comparing curves to lithology.",
        },
    )
    autofit(ws, max_width=34)
    apply_light_borders(ws)


def create_curve_samples(wb: Workbook) -> None:
    ws = wb.create_sheet("06_Curve_Samples_CSV")
    headers = [
        "project_code",
        "borehole_id",
        "depth_m",
        "gamma_api",
        "density_gcc",
        "resistivity_ohm_m",
        "caliper_mm",
        "sp_mv",
        "sonic_us_per_ft",
        "inclination_deg",
        "azimuth_deg",
        "other_curve_1_name",
        "other_curve_1_value",
        "source_file",
        "qa_flag",
    ]
    rows = [headers] + [["" for _ in headers] for _ in range(500)]
    add_table(ws, rows, "Curve_Samples_CSV")
    set_common_style(ws)
    add_validation(ws, "O2:O501", ["ok", "null_segment", "outlier", "depth_gap", "flatline", "needs_review"])
    add_comments(
        ws,
        {
            "A1": "Use this only if LAS cannot be provided. LAS is preferred for production.",
            "C1": "Depth must be numeric and sorted ascending per borehole.",
            "O1": "Flag known quality issues rather than deleting source samples.",
        },
    )
    autofit(ws, max_width=30)
    apply_light_borders(ws)


def create_seam_correlation(wb: Workbook) -> None:
    ws = wb.create_sheet("07_Seam_Correlation")
    headers = [
        "project_code",
        "block_name",
        "borehole_id",
        "section_line",
        "seam_name",
        "seam_alias_or_group",
        "top_depth_m",
        "base_depth_m",
        "thickness_m",
        "top_rl_m",
        "base_rl_m",
        "roof_lithology",
        "floor_lithology",
        "parting_count",
        "parting_total_thickness_m",
        "split_merge_flag",
        "fault_or_intrusion_note",
        "correlation_confidence",
        "evidence_sources",
        "neighbouring_boreholes_used",
        "picked_by",
        "picked_at",
        "approval_status",
        "export_marker_name",
        "remarks",
    ]
    rows = [headers]
    for i in range(2, 302):
        row = ["" for _ in headers]
        row[8] = f"=IF(OR(G{i}=\"\",H{i}=\"\"),\"\",H{i}-G{i})"
        rows.append(row)
    add_table(ws, rows, "Seam_Correlation")
    set_common_style(ws)
    add_validation(ws, "P2:P301", ["normal", "split", "merge", "missing", "fault_offset", "intrusion_affected", "uncertain"])
    add_validation(ws, "R2:R301", ["high", "medium", "low", "needs_review"])
    add_validation(ws, "W2:W301", ["draft", "reviewed", "approved", "rejected"])
    add_comments(
        ws,
        {
            "E1": "Use the approved seam dictionary. Put local names/aliases in column F.",
            "R1": "Confidence should reflect evidence quality, not just algorithm score.",
            "S1": "Examples: lithology log, LAS gamma/density, core image, coal quality, historical correlation chart.",
        },
    )
    autofit(ws, max_width=34)
    apply_light_borders(ws)


def create_core_images(wb: Workbook) -> None:
    ws = wb.create_sheet("08_Core_Images")
    headers = [
        "project_code",
        "borehole_id",
        "box_number",
        "tray_or_lane",
        "image_file_name",
        "original_folder_path",
        "from_depth_m",
        "to_depth_m",
        "depth_label_visible",
        "lane_count",
        "lane_order",
        "captured_by",
        "captured_at",
        "camera_or_device",
        "image_quality",
        "depth_mapping_status",
        "strip_or_crop_file",
        "ocr_depth_text",
        "annotation_status",
        "missing_core_notes",
        "fracture_or_defect_notes",
        "remarks",
    ]
    rows = [headers] + [["" for _ in headers] for _ in range(300)]
    add_table(ws, rows, "Core_Images")
    set_common_style(ws)
    add_validation(ws, "I2:I301", ["yes", "no", "partial", "unknown"])
    add_validation(ws, "O2:O301", ["good", "usable", "poor_lighting", "blurred", "needs_recapture"])
    add_validation(ws, "P2:P301", ["confirmed", "inferred", "pending", "conflict"])
    add_validation(ws, "S2:S301", ["none", "pending_geologist", "approved", "rejected"])
    add_comments(
        ws,
        {
            "G1": "Exact image-to-depth registration is essential before image AI is trusted.",
            "J1": "Capture 4-lane/5-lane tray pattern or other box geometry.",
            "P1": "Use pending/conflict rather than forcing uncertain mappings.",
        },
    )
    autofit(ws, max_width=34)
    apply_light_borders(ws)


def create_coal_quality(wb: Workbook) -> None:
    ws = wb.create_sheet("09_Coal_Quality")
    headers = [
        "project_code",
        "borehole_id",
        "sample_id",
        "parent_sample_id",
        "sample_type",
        "from_depth_m",
        "to_depth_m",
        "seam_name",
        "parting_or_band",
        "sample_basis",
        "lab_name",
        "lab_report_file",
        "analysis_date",
        "ash_pct",
        "moisture_pct",
        "volatile_matter_pct",
        "fixed_carbon_pct",
        "sulfur_pct",
        "gcv_kcal_kg",
        "ncv_kcal_kg",
        "hgi",
        "washability_available",
        "density_fraction",
        "qa_qc_status",
        "remarks",
    ]
    rows = [headers] + [["" for _ in headers] for _ in range(300)]
    add_table(ws, rows, "Coal_Quality")
    set_common_style(ws)
    add_validation(ws, "E2:E301", ["core", "channel", "composite", "ply", "ROM", "product", "other"])
    add_validation(ws, "J2:J301", ["adb", "db", "arb", "daf", "unknown"])
    add_validation(ws, "V2:V301", ["yes", "no", "partial"])
    add_validation(ws, "X2:X301", ["accepted", "repeat_required", "outlier", "rejected", "unknown"])
    add_comments(
        ws,
        {
            "J1": "Basis is mandatory before comparing quality values.",
            "N1": "Use numeric percent values, not text with percent symbols.",
            "S1": "State kcal/kg or provide alternate unit in remarks if different.",
        },
    )
    autofit(ws, max_width=34)
    apply_light_borders(ws)


def create_geotech_hydro_runtime(wb: Workbook) -> None:
    ws = wb.create_sheet("10_Geotech_Hydro_Runtime")
    headers = [
        "project_code",
        "borehole_id",
        "record_type",
        "from_depth_m",
        "to_depth_m",
        "timestamp",
        "parameter_name",
        "value",
        "unit",
        "measurement_method",
        "source_form_or_file",
        "submitted_by",
        "qa_status",
        "remarks",
    ]
    rows = [headers] + [["" for _ in headers] for _ in range(400)]
    add_table(ws, rows, "Geotech_Hydro_Runtime")
    set_common_style(ws)
    add_validation(ws, "C2:C401", ["RQD", "fracture", "point_load", "UCS", "water_level", "water_flow", "gas", "ROP", "torque", "bit_depth", "hole_depth", "field_observation", "other"])
    add_validation(ws, "M2:M401", ["raw", "reviewed", "approved", "rejected"])
    add_comments(
        ws,
        {
            "C1": "Flexible long-form table for parameters not stable enough to make first-class fields.",
            "D1": "Use depth range for interval observations; use timestamp for runtime observations.",
        },
    )
    autofit(ws, max_width=34)
    apply_light_borders(ws)


def create_mobile_forms(wb: Workbook) -> None:
    ws = wb.create_sheet("11_Mobile_Field_Forms")
    headers = [
        "form_name",
        "form_version",
        "field_key",
        "field_label",
        "data_type",
        "unit",
        "required",
        "dropdown_dictionary",
        "mobile_screen",
        "offline_allowed",
        "validation_rule",
        "canonical_mapping",
        "central_merge_policy",
        "photo_or_file_attachment",
        "remarks",
    ]
    rows = [
        headers,
        ["lithology_interval", "v1", "from_depth_m", "From depth", "number", "m", "yes", "", "Structured field log", "yes", ">=0", "lithology.from_depth", "replace_or_append_by_depth", "no", ""],
        ["lithology_interval", "v1", "to_depth_m", "To depth", "number", "m", "yes", "", "Structured field log", "yes", "> from_depth", "lithology.to_depth", "replace_or_append_by_depth", "no", ""],
        ["lithology_interval", "v1", "lithology_code", "Lithology code", "enum", "", "yes", "lithology_dictionary", "Structured field log", "yes", "must be approved or review status", "lithology.lithology_code", "field_stage_review", "no", ""],
        ["runtime", "v1", "hole_depth", "Hole depth", "number", "m", "no", "", "Runtime parameters", "yes", ">= current bit depth", "field_submission.payload.runtime_parameters", "append_observation", "no", ""],
        ["camera", "v1", "corebox_image", "Corebox image", "file", "", "no", "", "Camera upload", "yes", "jpg/png only", "source_file.corebox_image", "pending_depth_mapping", "yes", ""],
    ] + [["" for _ in headers] for _ in range(150)]
    add_table(ws, rows, "Mobile_Field_Forms")
    set_common_style(ws)
    add_validation(ws, "E2:E156", ["text", "number", "date", "datetime", "enum", "boolean", "file", "json"])
    add_validation(ws, "G2:G156", ["yes", "no", "conditional"])
    add_validation(ws, "J2:J156", ["yes", "no"])
    add_validation(ws, "N2:N156", ["yes", "no", "conditional"])
    add_comments(
        ws,
        {
            "L1": "Direct mapping to backend canonical fields or JSON payload.",
            "M1": "Examples: append_observation, replace_or_append_by_depth, pending_depth_mapping, conflict_review.",
        },
    )
    autofit(ws, max_width=34)
    apply_light_borders(ws)


def create_dictionaries(wb: Workbook) -> None:
    ws = wb.create_sheet("12_Dictionaries_Standards")
    headers = [
        "dictionary_name",
        "source_code_or_value",
        "approved_code",
        "approved_label",
        "category",
        "display_color_hex",
        "display_pattern",
        "status",
        "owner",
        "effective_from",
        "notes",
    ]
    rows = [
        headers,
        ["lithology", "COAL", "COAL", "Coal", "coal", "#202124", "solid_black", "approved", "Chief geologist", "", ""],
        ["lithology", "SH COAL", "SHCOAL", "Shaly Coal", "coal", "#30261F", "coal_shale_mix", "alias_review", "Chief geologist", "", "Observed in CTSJ workbook."],
        ["lithology", "CARB SHALE", "CARBSHL", "Carbonaceous Shale", "carbonaceous", "#4B5563", "shale_dark", "alias_review", "Chief geologist", "", ""],
        ["seam", "XI TOP (T)", "XI_TOP_T", "XI Top T", "seam", "", "", "needs_customer_approval", "Central geology", "", "Example from current sample data."],
        ["quality_basis", "adb", "ADB", "Air dried basis", "quality", "", "", "approved", "Quality team", "", ""],
    ] + [["" for _ in headers] for _ in range(300)]
    add_table(ws, rows, "Dictionaries_Standards")
    set_common_style(ws)
    add_validation(ws, "H2:H306", ["approved", "alias_review", "needs_customer_approval", "deprecated", "do_not_map"])
    add_comments(
        ws,
        {
            "A1": "Examples: lithology, seam, quality_basis, sample_type, approval_status, curve_alias, coordinate_system.",
            "B1": "Value exactly as it appears in Reliance source files.",
            "C1": "Approved canonical value used by GeoWorkbench and exports.",
        },
    )
    autofit(ws, max_width=34)
    apply_light_borders(ws)


def create_ai_corpus(wb: Workbook) -> None:
    ws = wb.create_sheet("13_AI_Learning_Corpus")
    headers = [
        "learning_item_id",
        "project_code",
        "borehole_id",
        "depth_from_m",
        "depth_to_m",
        "input_modalities_available",
        "target_label",
        "target_label_type",
        "approved_by",
        "approval_date",
        "source_files",
        "feature_notes",
        "train_validation_test_split",
        "data_quality_status",
        "permitted_use",
        "remarks",
    ]
    rows = [
        headers,
        ["AI-EX-001", "", "", "", "", "lithology, LAS, core_image, coal_quality", "COAL", "lithology_class", "", "", "", "Matched approved interval plus curve window.", "train", "example_only", "Reliance NDA internal", ""],
    ] + [["" for _ in headers] for _ in range(500)]
    add_table(ws, rows, "AI_Learning_Corpus")
    set_common_style(ws)
    add_validation(ws, "H2:H501", ["lithology_class", "coal_candidate", "seam_name", "seam_top_base", "quality_prediction", "image_class", "defect_class", "report_fact"])
    add_validation(ws, "M2:M501", ["train", "validation", "test", "holdout_customer_review"])
    add_validation(ws, "N2:N501", ["usable", "partial", "needs_review", "exclude"])
    add_validation(ws, "O2:O501", ["Reliance NDA internal", "anonymized_training_ok", "no_model_training", "review_required"])
    add_comments(
        ws,
        {
            "F1": "Examples: lithology, LAS, core_image, coal_quality, report_text, correction_history.",
            "M1": "Keep final test boreholes separate to measure real model performance.",
            "O1": "Explicitly capture allowed use for AI/model training under NDA.",
        },
    )
    autofit(ws, max_width=36)
    apply_light_borders(ws)


def create_reports_exports(wb: Workbook) -> None:
    ws = wb.create_sheet("14_Reports_Model_Exports")
    headers = [
        "artifact_id",
        "project_code",
        "block_name",
        "artifact_type",
        "artifact_name",
        "file_format",
        "software_source",
        "version_or_date",
        "related_boreholes_or_seams",
        "contains_tables",
        "contains_maps_or_sections",
        "editable_source_available",
        "target_workflow",
        "confidentiality",
        "notes",
    ]
    rows = [
        headers,
        ["REP-001", "", "", "final_geological_report", "", "PDF/DOCX", "", "", "", "yes", "yes", "preferred", "AI knowledge assistant/report drafting", "High", ""],
        ["EXP-001", "", "", "minex_import_template", "", "XLSX/CSV", "Minex or modelling package", "", "", "yes", "no", "yes", "Export Center finalization", "High", ""],
    ] + [["" for _ in headers] for _ in range(200)]
    add_table(ws, rows, "Reports_Model_Exports")
    set_common_style(ws)
    add_validation(ws, "D2:D203", ["final_geological_report", "litholog", "seam_correlation_chart", "section", "contour_plan", "seam_folio", "resource_statement", "model_export", "import_template", "dashboard_report", "other"])
    add_validation(ws, "J2:J203", ["yes", "no", "unknown"])
    add_validation(ws, "K2:K203", ["yes", "no", "unknown"])
    add_validation(ws, "L2:L203", ["yes", "no", "preferred", "unknown"])
    autofit(ws, max_width=36)
    apply_light_borders(ws)


def create_qa_rules(wb: Workbook) -> None:
    ws = wb.create_sheet("15_QA_Acceptance_Rules")
    rows = [
        ["Rule_ID", "Area", "Rule", "Severity", "Needed_dataset", "Why_it_matters", "Automation_status"],
        ["QA-001", "Intervals", "No negative or zero thickness intervals.", "Error", "GeoIntervals", "Blocks invalid logs and exports.", "Already implemented in backend validation"],
        ["QA-002", "Intervals", "No unexplained gaps or overlaps within reviewed depth coverage.", "Error/Warning", "GeoIntervals", "Required before correlation and modelling export.", "Already implemented"],
        ["QA-003", "Recovery/RQD", "Recovery cannot exceed interval/run thickness unless marked as source conflict.", "Warning", "GeoIntervals, Drilling Run", "Prevents bad geotechnical/recovery analytics.", "Already implemented"],
        ["QA-004", "RQD", "RQD values must state source format: fraction 0-1 or percent 0-100.", "Warning", "GeoIntervals, Dictionaries", "Current PBH/CTSJ examples differ.", "Needs data dictionary enforcement"],
        ["QA-005", "Lithology", "Unknown lithology codes must be preserved and flagged, not silently converted.", "Warning", "Dictionaries", "Protects geological meaning and AI labels.", "Partially implemented"],
        ["QA-006", "Seams", "Coal/carbonaceous intervals without seam names should be reviewable, not auto-filled.", "Info", "GeoIntervals, Seam Correlation", "Supports geologist-controlled seam naming.", "Already implemented"],
        ["QA-007", "Curves", "LAS borehole ID, depth range, null value, units, and curve aliases must be profiled.", "Warning", "Curve Metadata", "Prevents misleading curve/lithology comparisons.", "Partially implemented"],
        ["QA-008", "Curves", "Document any depth shift before using logs for interval adjustment.", "Warning", "Curve Metadata, GeoIntervals", "Depth mismatch is a common source of wrong seam picks.", "Needs workflow support"],
        ["QA-009", "Core Images", "Every image used as evidence needs confirmed or explicitly inferred depth mapping.", "Warning", "Core Images", "Image AI and visual evidence are unreliable without it.", "Partially implemented"],
        ["QA-010", "GIS", "Every spatial layer must declare CRS/EPSG or local grid transform.", "Error", "Spatial/GIS", "Prevents wrong maps, sections, and mobile positions.", "New requirement"],
        ["QA-011", "AI", "Training labels must be approved final interpretations or explicitly tagged as raw/draft.", "Error", "AI Learning Corpus", "Avoids training on unapproved geology.", "New requirement"],
        ["QA-012", "Exports", "Customer target template must accept generated export without manual column reshaping.", "Error", "Import/Export Templates", "Defines production readiness for Reliance.", "New requirement"],
    ]
    add_table(ws, rows, "QA_Acceptance_Rules", COLORS["header_alt"])
    set_common_style(ws)
    add_validation(ws, "D2:D200", ["Error", "Warning", "Info", "Error/Warning"])
    autofit(ws, max_width=44)
    apply_light_borders(ws)


def create_sources(wb: Workbook) -> None:
    ws = wb.create_sheet("16_Source_References")
    rows = [
        ["Source", "URL", "How_used_in_this_workbook"],
        ["ACARP CoalLog Manual abstract", "https://www.acarp.com.au/abstracts.aspx?repId=C21003", "Supports request for coal borehole logging sheets, dictionaries, and transfer format concepts."],
        ["AusIMM/ACARP CoalLog Geology and Geotechnical Training Manual", "https://www.ausimm.com/globalassets/coallog/coallog-v2.0-training-manual.pdf", "Supports geophysical logging, drilling, lithology, recovery, geotechnical, hydrogeology, and sampling data requirements."],
        ["USGS LAS Format", "https://www.usgs.gov/programs/national-geological-and-geophysical-data-preservation-program/las-format", "Supports LAS as preferred raw geophysical log exchange format and required sections."],
        ["USGS Circular 891: Geophysical Logs as a Source of Coal Bed Data", "https://pubs.usgs.gov/circ/c891/geophysical.htm", "Supports geophysical logs for coal bed thickness, depth, and correlation context."],
        ["OGC GeoPackage standard", "https://www.ogc.org/standards/geopackage/", "Supports GeoPackage as portable geospatial/offline-mobile exchange format."],
        ["PIB Ministry of Mines release on exploration data and UNFC/MEMC", "https://www.pib.gov.in/PressReleasePage.aspx?PRID=1744880", "Supports Indian exploration/reporting context and need to preserve geological reports and resource data."],
        ["MECL Mineral Exploration page", "https://mecl.co.in/ContentPageMecl.aspx?Antispam=5986bcdc-bac7-4286-8f12-6f7e9bac0700&ControlID=4&Lng=EN&MyAntispam=a3782f6e-6c86-401c-b726-e2951ccb487d&page=mineral-exploration", "Supports request for borehole logs, lithologs, seam correlation charts, geophysical interpretation reports, analytical reports, maps, sections, and resource statements."],
        ["Applications of Geophysical Logs to Coal Mining", "https://www.mdpi.com/2079-9276/9/2/11", "Supports advanced geophysical-log applications in coal mining such as geotechnical characterization and derived analytics."],
        ["GeoWorkbench README", "D:/Source/geoworkbench/README.md", "Existing product direction: import, visualization, validation, AI suggestions, reporting, and feedback capture."],
        ["GeoWorkbench canonical data model strategy", "D:/Source/geoworkbench/docs/canonical-data-model-and-exchange-strategy.md", "Existing canonical groups: boreholes, intervals, curves, images, source/provenance, mobile submissions, import/export profiles."],
        ["GeoWorkbench backend models", "D:/Source/geoworkbench/backend/app/db/models.py", "Current database entities that these request templates feed."],
    ]
    add_table(ws, rows, "Source_References", COLORS["header_alt"])
    set_common_style(ws)
    autofit(ws, max_width=80)
    apply_light_borders(ws)


def build_workbook() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    create_readme(wb)
    create_dataset_request(wb)
    create_delivery_plan(wb)
    create_borehole_master(wb)
    create_geointervals(wb)
    create_curve_metadata(wb)
    create_curve_samples(wb)
    create_seam_correlation(wb)
    create_core_images(wb)
    create_coal_quality(wb)
    create_geotech_hydro_runtime(wb)
    create_mobile_forms(wb)
    create_dictionaries(wb)
    create_ai_corpus(wb)
    create_reports_exports(wb)
    create_qa_rules(wb)
    create_sources(wb)

    for ws in wb.worksheets:
        ws.sheet_format.defaultColWidth = 14
        ws.sheet_format.defaultRowHeight = 20

    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


def verify_workbook(path: Path) -> dict:
    wb = load_workbook(path, data_only=False)
    required = [
        "00_ReadMe",
        "01_Dataset_Request",
        "04_GeoIntervals",
        "05_Curve_Metadata",
        "06_Curve_Samples_CSV",
        "07_Seam_Correlation",
        "13_AI_Learning_Corpus",
        "16_Source_References",
    ]
    missing = [sheet for sheet in required if sheet not in wb.sheetnames]
    formula_count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
    return {
        "path": str(path),
        "sheet_count": len(wb.sheetnames),
        "missing_required_sheets": missing,
        "formula_count": formula_count,
        "dataset_request_rows": wb["01_Dataset_Request"].max_row - 1,
        "geo_interval_template_rows": wb["04_GeoIntervals"].max_row - 1,
    }


if __name__ == "__main__":
    output = build_workbook()
    print(verify_workbook(output))
