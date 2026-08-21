import re
from copy import deepcopy
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload
from sqlalchemy.orm.attributes import flag_modified

from app.db.models import (
    Borehole,
    DisplayLayout,
    LithologyInterval,
    Project,
    SeamInterval,
    Site,
    SourceImport,
)
from app.domains.display_layouts.defaults import default_borehole_layout
from app.services.excel_import import normalize_lithology
from app.services.las_import import import_las_curves, profile_las_file
from app.services.data_stage import IMPORTED_INTERPRETED, merge_stage_metadata
from app.services.validation.borehole_validation import replace_validation_issues, validate_borehole


PROJECT_CODE = "RELIANCE-COAL"
PROJECT_NAME = "Reliance Coal Data"
SITE_CODE = "MGCA"
SITE_NAME = "MGCA Borehole Set"


def _clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _num(value: Any) -> float | None:
    value = _clean(value)
    if value in {None, "-", ""}:
        return None
    try:
        return round(float(value), 4)
    except (TypeError, ValueError):
        return None


def _text(value: Any) -> str | None:
    value = _clean(value)
    if value is None:
        return None
    return str(value)


def normalize_borehole_code(value: Any) -> str | None:
    text = _text(value)
    if not text:
        return None
    text = text.upper().replace(" ", "")
    match = re.match(r"^(MGCA)-0?(\d+)$", text)
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}"
    return re.sub(r"[^A-Z0-9_-]+", "-", text).strip("-")


def _rows(path: Path, *, header_row: int = 1) -> tuple[list[str], list[tuple[int, tuple[Any, ...]]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header: list[str] | None = None
    records: list[tuple[int, tuple[Any, ...]]] = []
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_number < header_row:
            continue
        if row_number == header_row:
            header = [str(value).strip() if value is not None else "" for value in row]
            continue
        if not row or all(_clean(value) is None for value in row):
            continue
        records.append((row_number, row))
    if header is None:
        raise ValueError(f"Could not read header row from {path.name}")
    return header, records


def _index(header: list[str], name: str) -> int:
    try:
        return header.index(name)
    except ValueError as exc:
        raise ValueError(f"Missing expected column '{name}'") from exc


def _value(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def read_collar(path: Path) -> dict[str, dict]:
    header, rows = _rows(path, header_row=2)
    bh_idx = _index(header, "BH No.")
    return {
        code: {
            "coalgrid_northing": _num(_value(row, 1)),
            "coalgrid_easting": _num(_value(row, 2)),
            "utm_easting": _num(_value(row, 3)),
            "utm_northing": _num(_value(row, 4)),
            "source_row": row_number,
        }
        for row_number, row in rows
        if (code := normalize_borehole_code(_value(row, bh_idx)))
    }


def read_lithology(path: Path) -> dict[str, list[dict]]:
    header, rows = _rows(path)
    col = {name: _index(header, name) for name in header if name}
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row_number, row in rows:
        code = normalize_borehole_code(_value(row, col["Borehole No."]))
        from_depth = _num(_value(row, col["From (m)"]))
        to_depth = _num(_value(row, col["To (m)"]))
        lithology_source = _text(_value(row, col["Lithology"]))
        if not code or from_depth is None or to_depth is None or not lithology_source:
            continue
        normalized = normalize_lithology(lithology_source)
        grouped[code].append(
            {
                "source_row": row_number,
                "from_depth": from_depth,
                "to_depth": to_depth,
                "thickness": _num(_value(row, col["Total Length (m)"])),
                "recovery": _num(_value(row, col["Total Recovery (m)"])),
                "recovery_percent": _num(_value(row, col["Recovery %"])),
                "recovery_thickness": _num(_value(row, col["Recovery Thickness (m)"])),
                "extrapolated_thickness": _num(_value(row, col["Extrapolated Thickness (m)"])),
                "floor_depth": _num(_value(row, col["Floor Depth (m)"])),
                "lithology_source": lithology_source,
                "lithology_code": normalized["code"],
                "lithology_label": normalized["label"],
                "display_color": normalized["color"],
                "logged_color": _text(_value(row, col["Colour"])),
                "structural_features": _text(_value(row, col["Structural Features"])),
                "other_details": _text(_value(row, col["Other Details"])),
                "formation": _text(_value(row, col["Formation"])),
                "seam_name": _text(_value(row, col["Seam Name"])),
            }
        )
    return grouped


def read_band_by_band(path: Path) -> dict[tuple[str, str, float, float], list[dict]]:
    header, rows = _rows(path)
    col = {name: _index(header, name) for name in header if name}
    grouped: dict[tuple[str, str, float, float], list[dict]] = defaultdict(list)
    for row_number, row in rows:
        code = normalize_borehole_code(_value(row, col["Borehole No."]))
        seam = _text(_value(row, col["Seam Name"]))
        from_depth = _num(_value(row, col["Depth From (m)"]))
        to_depth = _num(_value(row, col["Depth To (m)"]))
        if not code or not seam or from_depth is None or to_depth is None:
            continue
        grouped[(code, seam, from_depth, to_depth)].append(
            {
                "source_row": row_number,
                "sample_no": _text(_value(row, col["Sample No"])),
                "lithology": _text(_value(row, col["Lithology"])),
                "other_details": _text(_value(row, col["Other Details"])),
                "weight_gms": _num(_value(row, col["Wt (gms)"])),
                "moisture_percent": _num(_value(row, col["M%"])),
                "ash_percent": _num(_value(row, col["Ash%"])),
                "vm_percent": _num(_value(row, col["VM%"])),
                "fc_percent": _num(_value(row, col["FC%"])),
                "rf": _text(_value(row, col["R/F"])) if "R/F" in col else None,
            }
        )
    return grouped


def read_overall_analysis(path: Path) -> dict[tuple[str, str, float, float], list[dict]]:
    header, rows = _rows(path)
    col = {name: _index(header, name) for name in header if name}
    grouped: dict[tuple[str, str, float, float], list[dict]] = defaultdict(list)
    for row_number, row in rows:
        code = normalize_borehole_code(_value(row, col["Borehole No."]))
        seam = _text(_value(row, col["Seam Name"]))
        from_depth = _num(_value(row, col["Seam Depth From (m)"]))
        to_depth = _num(_value(row, col["Seam Depth To (m)"]))
        if not code or not seam or from_depth is None or to_depth is None:
            continue
        grouped[(code, seam, from_depth, to_depth)].append(
            {
                "source_row": row_number,
                "mode": _text(_value(row, col["Mode of Analysis"])),
                "analysed_thickness": _num(_value(row, col["Analysed Thickness"])),
                "moisture_percent": _num(_value(row, col["M%"])),
                "ash_percent": _num(_value(row, col["Ash%"])),
                "vm_percent": _num(_value(row, col["VM%"])),
                "fc_percent": _num(_value(row, col["FC%"])),
                "uvm_percent": _num(_value(row, col["UVM%"])),
                "cv_kcal_kg": _num(_value(row, col["CV Kcal/kg"])),
                "ucv": _num(_value(row, col["UCV"])),
                "gcv_kcal_kg": _num(_value(row, col["GCV Kcal/kg"])),
                "grade": _text(_value(row, col["Grade"])),
                "determination": _text(_value(row, col["Determined/ Calculated"])),
                "seam_or_section": _text(_value(row, col["Seam/ Section"])),
                "remarks": _text(_value(row, col["Remarks"])) if "Remarks" in col else None,
            }
        )
    return grouped


def read_dirt_bands(path: Path) -> dict[str, list[dict]]:
    header, rows = _rows(path)
    col = {name: _index(header, name) for name in header if name}
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row_number, row in rows:
        code = normalize_borehole_code(_value(row, col["Borehole No"]))
        seam = _text(_value(row, col["Seam Name"]))
        from_depth = _num(_value(row, col["Depth of Seam - From (m)"]))
        to_depth = _num(_value(row, col["Depth of Seam - To (m)"]))
        if not code or not seam or from_depth is None or to_depth is None:
            continue
        grouped[code].append(
            {
                "source_row": row_number,
                "name": seam,
                "from_depth": from_depth,
                "to_depth": to_depth,
                "thickness": _num(_value(row, col["Thickness (m)"])),
                "csh_high_bands_no": _num(_value(row, col["CSH HIGH Bands - No"])),
                "csh_high_bands_total_thick": _num(_value(row, col["CSH HIGH Bands - Total Thick (m)"])),
                "csh_high_bands_percent": _num(_value(row, col["CSH HIGH Bands - (%) of Seam"])),
                "ob_bands_no": _num(_value(row, col["OB Bands - No"])),
                "ob_bands_total_thick": _num(_value(row, col["OB Bands - Total Thick (m)"])),
                "ob_bands_percent": _num(_value(row, col["OB Bands - (%) of Seam"])),
                "combined_bands_no": _num(_value(row, col["Total CSH HIGH+OB - No"])),
                "combined_bands_total_thick": _num(_value(row, col["Total CSH HIGH+OB - Total Thick (m)"])),
                "combined_bands_percent": _num(_value(row, col["Total CSH HIGH+OB - (%) of Seam"])),
                "csh_ob_gt_030_thickness": _num(_value(row, col["CSH HIGH & OB (>0.30m) Thickness (m)"])),
                "bcs_thickness": _num(_value(row, col["BCS Thickness (m)"])),
                "i30_thickness": _num(_value(row, col["I-30 Thickness (m)"])),
                "i100_thickness": _num(_value(row, col["I-100 Thickness (m)"])),
            }
        )
    return grouped


def _ensure_project_site(db: Session) -> Site:
    project = db.scalar(select(Project).where(Project.code == PROJECT_CODE))
    if project is None:
        project = Project(code=PROJECT_CODE, name=PROJECT_NAME)
        db.add(project)
        db.flush()
    site = db.scalar(select(Site).where(Site.project_id == project.id).where(Site.code == SITE_CODE))
    if site is None:
        site = Site(project=project, code=SITE_CODE, name=SITE_NAME)
        db.add(site)
        db.flush()
    return site


def _delete_existing_borehole(db: Session, code: str) -> None:
    existing = db.scalar(
        select(Borehole)
        .where(Borehole.code == code)
        .options(
            selectinload(Borehole.lithology_intervals),
            selectinload(Borehole.seam_intervals),
            selectinload(Borehole.curves),
            selectinload(Borehole.core_images),
            selectinload(Borehole.display_layouts),
            selectinload(Borehole.validation_issues),
            selectinload(Borehole.ai_suggestions),
            selectinload(Borehole.source_imports),
            selectinload(Borehole.source_files),
            selectinload(Borehole.field_submissions),
            selectinload(Borehole.export_jobs),
        )
    )
    if existing is not None:
        db.delete(existing)
        db.flush()


def _configure_layout(layout: dict, curve_summaries: list[dict], *, has_core_images: bool = False) -> dict:
    tracks = layout["widgets"]["log-widget"]["tracks"]
    for track in tracks:
        if track["id"] == "core-images":
            track["visible"] = True
        if track["id"] == "rqd":
            track["visible"] = False
        if track["id"] == "curves":
            track["curves"] = [
                {
                    "curveKey": curve["key"],
                    "label": curve["label"],
                    "unit": curve["unit"],
                    "color": curve.get("color") or "#64748b",
                    "visible": True,
                    "scale": {
                        "mode": "manual",
                        "min": curve["min"] if curve.get("min") is not None else 0,
                        "max": curve["max"] if curve.get("max") is not None else 1,
                    },
                    "normalization": {"enabled": True, "method": "linear-track-scale"},
                }
                for curve in curve_summaries
            ]
    return layout


def import_reliance_dataset(db: Session, data_root: Path, las_root: Path | None = None) -> dict:
    collar = read_collar(data_root / "Collar_10BH.xlsx")
    lithology = read_lithology(data_root / "Lithology_10BH.xlsx")
    dirt_bands = read_dirt_bands(data_root / "Dirt_band_10BH.xlsx")
    band_rows = read_band_by_band(data_root / "Band_by_band_10BH.xlsx")
    analysis_rows = read_overall_analysis(data_root / "Overall analysis_10BH.xlsx")
    site = _ensure_project_site(db)

    imported = []
    for code in sorted(lithology):
        _delete_existing_borehole(db, code)
        intervals = sorted(lithology[code], key=lambda item: (item["from_depth"], item["to_depth"]))
        total_depth = max((item["to_depth"] for item in intervals), default=0)
        borehole = Borehole(
            site=site,
            code=code,
            title=f"{code} Reliance Borehole",
            state="Reliance real data",
            total_depth=total_depth,
            source_workbook="RelianceData/Data_10BH.zip",
            source_sheet="Lithology_10BH.xlsx",
            workflow_status="imported_for_central_review",
            attributes={
                "project": PROJECT_CODE,
                "block": SITE_CODE,
                "collar": collar.get(code),
                "source_package": "RelianceData",
            },
        )
        for item in intervals:
            borehole.lithology_intervals.append(
                LithologyInterval(
                    id=f"{code.lower()}-lith-{item['source_row']}",
                    source_row=item["source_row"],
                    from_depth=item["from_depth"],
                    to_depth=item["to_depth"],
                    lithology_code=item["lithology_code"],
                    lithology_label=item["lithology_label"],
                    display_color=item["display_color"],
                    logged_color=item.get("logged_color"),
                    seam_name=item.get("seam_name"),
                    recovery=item.get("recovery"),
                    recovery_percent=item.get("recovery_percent"),
                    rqd=None,
                    structural_features=item.get("structural_features"),
                    remark=item.get("other_details"),
                    attributes=merge_stage_metadata(
                        {
                            "formation": item.get("formation"),
                            "floor_depth": item.get("floor_depth"),
                            "recovery_thickness": item.get("recovery_thickness"),
                            "extrapolated_thickness": item.get("extrapolated_thickness"),
                            "lithology_source": item.get("lithology_source"),
                        },
                        IMPORTED_INTERPRETED,
                        source_type="reliance_consolidated_excel",
                        source_name="RelianceData/Data_10BH.zip",
                    ),
                )
            )

        for item in sorted(dirt_bands.get(code, []), key=lambda row: (row["from_depth"], row["to_depth"], row["name"])):
            key = (code, item["name"], item["from_depth"], item["to_depth"])
            borehole.seam_intervals.append(
                SeamInterval(
                    id=f"{code.lower()}-seam-{item['source_row']}",
                    source_row=item["source_row"],
                    name=item["name"],
                    from_depth=item["from_depth"],
                    to_depth=item["to_depth"],
                    thickness=item.get("thickness"),
                    lithology_code=None,
                    lithology_label=None,
                    attributes=merge_stage_metadata(
                        {
                            **{k: v for k, v in item.items() if k not in {"source_row", "name", "from_depth", "to_depth", "thickness"}},
                            "band_by_band": band_rows.get(key, []),
                            "overall_analysis": analysis_rows.get(key, []),
                        },
                        IMPORTED_INTERPRETED,
                        source_type="reliance_consolidated_excel",
                        source_name="RelianceData/Data_10BH.zip",
                    ),
                )
            )

        borehole.source_imports.append(
            SourceImport(
                import_type="excel",
                source_name="RelianceData/Data_10BH.zip",
                status="imported",
                summary={
                    "parser": "reliance_consolidated_excel_v1",
                    "metadata": {"block": SITE_CODE, "collar": collar.get(code)},
                    "files": [
                        "Collar_10BH.xlsx",
                        "Lithology_10BH.xlsx",
                        "Band_by_band_10BH.xlsx",
                        "Dirt_band_10BH.xlsx",
                        "Overall analysis_10BH.xlsx",
                    ],
                    "counts": {
                        "lithology_intervals": len(borehole.lithology_intervals),
                        "seam_intervals": len(borehole.seam_intervals),
                    },
                },
            )
        )
        borehole.display_layouts.append(
            DisplayLayout(name="Reliance Review Display", mode="runtime", settings=default_borehole_layout())
        )
        db.add(borehole)
        db.flush()
        replace_validation_issues(borehole, validate_borehole(borehole))
        db.commit()
        db.refresh(borehole)

        las_summary = None
        if las_root is not None:
            las_path = next((path for path in [las_root / f"{code}.las", las_root / f"{code.replace('-0', '-')}.las"] if path.exists()), None)
            if las_path is not None:
                las_summary = import_las_curves(db, borehole, las_path, replace_existing=True)
                db.refresh(borehole, attribute_names=["display_layouts"])
                layout = borehole.display_layouts[0]
                layout.settings = _configure_layout(deepcopy(layout.settings), las_summary["curves"])
                flag_modified(layout, "settings")
                db.add(layout)
                db.commit()

        imported.append(
            {
                "code": code,
                "borehole_id": borehole.id,
                "lithology_intervals": len(intervals),
                "seam_intervals": len(dirt_bands.get(code, [])),
                "total_depth": total_depth,
                "collar": collar.get(code),
                "las": las_summary,
            }
        )

    return {
        "project": PROJECT_CODE,
        "site": SITE_CODE,
        "boreholes": imported,
        "source": {
            "data_root": str(data_root),
            "las_root": str(las_root) if las_root else None,
        },
    }


def profile_reliance_dataset(data_root: Path, las_root: Path | None = None) -> dict:
    collar = read_collar(data_root / "Collar_10BH.xlsx")
    lithology = read_lithology(data_root / "Lithology_10BH.xlsx")
    dirt_bands = read_dirt_bands(data_root / "Dirt_band_10BH.xlsx")
    result = {
        "borehole_count": len(lithology),
        "boreholes": [
            {
                "code": code,
                "collar": collar.get(code),
                "lithology_intervals": len(rows),
                "seam_intervals": len(dirt_bands.get(code, [])),
                "depth_from": min((row["from_depth"] for row in rows), default=None),
                "depth_to": max((row["to_depth"] for row in rows), default=None),
            }
            for code, rows in sorted(lithology.items())
        ],
    }
    if las_root is not None:
        result["las"] = [
            {
                "file": path.name,
                "borehole_code": normalize_borehole_code(path.stem),
                "profile": profile_las_file(path),
            }
            for path in sorted(las_root.glob("*.las"))
        ]
    return result
